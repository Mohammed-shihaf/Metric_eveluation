"""Export comparison results to Excel."""

from __future__ import print_function

import io
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import PatternFill

MATCH_FILL = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
MISMATCH_FILL = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")

DIFFERENCE_HEADERS = [
    "branch",
    "mismatched_field",
    "S3 Data",
    "Local Data",
    "S3 vs Local",
    "delta",
    "S3 tool executed",
    "Local tool executed",
]


def _sheet_from_rows(ws, headers, rows):
    ws.append(headers)
    for row in rows:
        ws.append([row.get(h, "") for h in headers])


def _tool_exec_label(executed, tool_name, executed_tool=""):
    tool = (executed_tool or tool_name or "").strip()
    if executed is True or str(executed).lower() in ("true", "1", "yes"):
        return "Yes — %s" % tool if tool else "Yes"
    if executed is False or str(executed).lower() in ("false", "0", "no"):
        return "No"
    if tool:
        return tool
    return "—"


def _difference_rows(result, mismatches_only=False):
    """Explicit S3-vs-local diff rows for Matches/Mismatches sheets."""
    branch = result.get("branch_name", "")
    pair = result.get("s3_vs_local") or {}
    s3_tool = _tool_exec_label(
        result.get("s3_executed"),
        result.get("s3_tool_name", ""),
        result.get("s3_executed_tool", ""),
    )
    local_tool = _tool_exec_label(
        result.get("local_real_tool"),
        result.get("local_tool_name", ""),
        result.get("local_executed_tool", ""),
    )
    rows = []
    for diff in (pair.get("field_diffs") or []) + (pair.get("metric_diffs") or []):
        is_match = diff.get("match") is True
        if mismatches_only and is_match:
            continue
        if not mismatches_only and diff.get("match") is False:
            continue
        rows.append({
            "branch": branch,
            "mismatched_field": diff.get("field", ""),
            "S3 Data": diff.get("s3", ""),
            "Local Data": diff.get("local", ""),
            "S3 vs Local": "MATCH" if is_match else "MISMATCH",
            "delta": diff.get("delta", ""),
            "S3 tool executed": s3_tool,
            "Local tool executed": local_tool,
        })
    return rows


def _summary_mismatch_detail(result):
    """Per-branch joined mismatch summary for the Summary sheet."""
    fields = []
    s3_parts = []
    local_parts = []
    for row in _difference_rows(result, mismatches_only=True):
        field = row.get("mismatched_field", "")
        if field:
            fields.append(field)
            s3_parts.append("%s=%s" % (field, row.get("S3 Data", "")))
            local_parts.append("%s=%s" % (field, row.get("Local Data", "")))
    return ", ".join(fields), "; ".join(s3_parts), "; ".join(local_parts)


def _flatten_diffs(branch, pair_name, pair):
    rows = []
    if not pair:
        return rows
    for diff in (pair.get("field_diffs") or []) + (pair.get("metric_diffs") or []):
        row = {
            "branch": branch,
            "pair": pair_name,
            "field": diff.get("field", ""),
            "match": diff.get("match"),
            "delta": diff.get("delta", ""),
        }
        for key, val in diff.items():
            if key not in ("field", "match", "delta"):
                row[key] = val
        rows.append(row)
    return rows


def _count_pair_diffs(pair):
    matched = mismatched = 0
    if not pair:
        return matched, mismatched
    seen = set()
    for diff in (pair.get("field_diffs") or []) + (pair.get("metric_diffs") or []):
        key = (diff.get("field"), diff.get("match"))
        if key in seen:
            continue
        seen.add(key)
        if diff.get("match"):
            matched += 1
        else:
            mismatched += 1
    return matched, mismatched


def _local_result_rows(results):
    rows = []
    for row in results or []:
        metrics = row.get("local_metric_values") or {}
        base = {
            "branch": row.get("branch_name", ""),
            "local_status": row.get("local_status", ""),
            "tool_name": row.get("local_tool_name", ""),
            "executed_tool": row.get("local_executed_tool", ""),
            "real_tool": row.get("local_real_tool", ""),
            "raw_summary": row.get("local_raw_summary", ""),
        }
        if not metrics:
            rows.append(base)
            continue
        for key, val in sorted(metrics.items()):
            item = dict(base)
            item["metric"] = key
            item["metric_value"] = val
            rows.append(item)
    return rows


def _fill_data_rows(ws, headers, rows, fill, match_col="S3 vs Local"):
    if match_col not in headers:
        return
    col_idx = headers.index(match_col) + 1
    for row_idx, row in enumerate(rows, start=2):
        if fill or row.get(match_col) == "MISMATCH":
            row_fill = MISMATCH_FILL if row.get(match_col) == "MISMATCH" else fill
            if row_fill:
                for c in range(1, len(headers) + 1):
                    ws.cell(row=row_idx, column=c).fill = row_fill


def build_comparison_workbook(results, whitebox_by_branch=None, path=None):
    """Build multi-sheet xlsx from comparison result dicts."""
    wb = Workbook()
    wb.remove(wb.active)
    whitebox_by_branch = whitebox_by_branch or {}

    summary_ws = wb.create_sheet("Summary")
    summary_headers = [
        "branch", "verdict", "taxonomy_status", "taxonomy_vs_s3", "taxonomy_vs_local",
        "s3_status", "local_status", "sonar_status",
        "s3_vs_local_verdict", "local_vs_sonar_verdict",
        "matched_fields", "mismatched_fields",
        "mismatched_fields_detail", "S3 Data", "Local Data",
        "S3 tool executed", "Local tool executed",
        "summary",
    ]
    summary_rows = []
    s3_local_rows = []
    local_sonar_rows = []
    status_rows = []
    failing_rows = []
    match_rows = []
    mismatch_rows = []
    local_result_rows = []

    for row in results or []:
        branch = row.get("branch_name", "")
        wb_info = whitebox_by_branch.get(branch, {})
        s3_local = row.get("s3_vs_local") or {}
        local_sonar = row.get("local_vs_sonar") or {}
        matched, mismatched = _count_pair_diffs(s3_local)
        m2, mm2 = _count_pair_diffs(local_sonar)
        matched += m2
        mismatched += mm2
        detail_fields, detail_s3, detail_local = _summary_mismatch_detail(row)
        s3_tool = _tool_exec_label(
            row.get("s3_executed"),
            row.get("s3_tool_name", ""),
            row.get("s3_executed_tool", ""),
        )
        local_tool = _tool_exec_label(
            row.get("local_real_tool"),
            row.get("local_tool_name", ""),
            row.get("local_executed_tool", ""),
        )
        summary_rows.append({
            "branch": branch,
            "verdict": row.get("verdict", ""),
            "taxonomy_status": row.get("taxonomy_status", ""),
            "taxonomy_vs_s3": row.get("taxonomy_vs_s3", ""),
            "taxonomy_vs_local": row.get("taxonomy_vs_local", ""),
            "s3_status": row.get("s3_status", ""),
            "local_status": row.get("local_status", ""),
            "sonar_status": row.get("sonar_status", ""),
            "s3_vs_local_verdict": s3_local.get("verdict", ""),
            "local_vs_sonar_verdict": local_sonar.get("verdict", ""),
            "matched_fields": matched,
            "mismatched_fields": mismatched,
            "mismatched_fields_detail": detail_fields,
            "S3 Data": detail_s3,
            "Local Data": detail_local,
            "S3 tool executed": s3_tool,
            "Local tool executed": local_tool,
            "summary": row.get("summary", ""),
        })
        status_rows.append({
            "branch": branch,
            "s3_status": row.get("s3_status", ""),
            "local_status": row.get("local_status", ""),
            "sonar_status": row.get("sonar_status", ""),
            "taxonomy_status_ref": row.get("taxonomy_status", ""),
            "taxonomy_vs_s3": row.get("taxonomy_vs_s3", ""),
            "taxonomy_vs_local": row.get("taxonomy_vs_local", ""),
            "run_status": wb_info.get("run_status", ""),
            "tasks": "%s/%s" % (
                wb_info.get("completed_tasks", 0), wb_info.get("total_tasks", 0),
            ) if wb_info.get("total_tasks") else "",
            "failed_tasks": wb_info.get("failed_tasks", ""),
            "run_health": wb_info.get("run_health", ""),
        })
        branch_s3_local = _flatten_diffs(branch, "s3_vs_local", s3_local)
        branch_local_sonar = _flatten_diffs(branch, "local_vs_sonar", local_sonar)
        s3_local_rows.extend(branch_s3_local)
        local_sonar_rows.extend(branch_local_sonar)
        match_rows.extend(_difference_rows(row, mismatches_only=False))
        mismatch_rows.extend(_difference_rows(row, mismatches_only=True))
        local_result_rows.extend(_local_result_rows([row]))
        for sec in wb_info.get("failing_sections") or []:
            failing_rows.append({
                "branch": branch,
                "testing_type": sec.get("testing_type", ""),
                "technique": sec.get("technique", ""),
                "classification": sec.get("classification", ""),
                "normalized_score": sec.get("normalized_score", ""),
            })

    _sheet_from_rows(summary_ws, summary_headers, summary_rows)
    verdict_col = summary_headers.index("verdict") + 1
    detail_col = summary_headers.index("mismatched_fields_detail") + 1
    for row_idx, row in enumerate(summary_rows, start=2):
        fill = None
        if row.get("verdict") == "MATCH":
            fill = MATCH_FILL
        elif row.get("verdict") == "MISMATCH":
            fill = MISMATCH_FILL
        if fill:
            summary_ws.cell(row=row_idx, column=verdict_col).fill = fill
        if row.get("mismatched_fields_detail"):
            summary_ws.cell(row=row_idx, column=detail_col).fill = MISMATCH_FILL

    matches_ws = wb.create_sheet("Matches")
    _sheet_from_rows(matches_ws, DIFFERENCE_HEADERS, match_rows)
    _fill_data_rows(matches_ws, DIFFERENCE_HEADERS, match_rows, MATCH_FILL)

    mismatches_ws = wb.create_sheet("Mismatches")
    _sheet_from_rows(mismatches_ws, DIFFERENCE_HEADERS, mismatch_rows)
    _fill_data_rows(mismatches_ws, DIFFERENCE_HEADERS, mismatch_rows, MISMATCH_FILL)

    local_ws = wb.create_sheet("Local results")
    local_headers = sorted({k for r in local_result_rows for k in r}) if local_result_rows else [
        "branch", "local_status", "tool_name", "executed_tool", "real_tool", "metric", "metric_value",
    ]
    _sheet_from_rows(local_ws, local_headers, local_result_rows)

    status_ws = wb.create_sheet("Report status")
    _sheet_from_rows(
        status_ws,
        [
            "branch", "s3_status", "local_status", "sonar_status",
            "taxonomy_status_ref", "taxonomy_vs_s3", "taxonomy_vs_local",
            "run_status", "tasks", "failed_tasks", "run_health",
        ],
        status_rows,
    )

    s3_local_ws = wb.create_sheet("S3 vs Local")
    s3_headers = sorted({k for r in s3_local_rows for k in r}) if s3_local_rows else [
        "branch", "pair", "field", "match", "delta",
    ]
    _sheet_from_rows(s3_local_ws, s3_headers, s3_local_rows)

    ls_ws = wb.create_sheet("Local vs Sonar")
    ls_headers = sorted({k for r in local_sonar_rows for k in r}) if local_sonar_rows else [
        "branch", "pair", "field", "match", "delta",
    ]
    _sheet_from_rows(ls_ws, ls_headers, local_sonar_rows)

    fail_ws = wb.create_sheet("Failing sections")
    _sheet_from_rows(
        fail_ws,
        ["branch", "testing_type", "technique", "classification", "normalized_score"],
        failing_rows,
    )

    if path:
        wb.save(str(path))
        return str(path)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.getvalue()
