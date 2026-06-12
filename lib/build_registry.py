#!/usr/bin/env python3
"""Build config/metrics_registry.yaml from Book.xlsx."""

from __future__ import print_function

import os
import re
import sys

try:
    import yaml
except ImportError:
    yaml = None

import openpyxl

ROOT = os.path.abspath(os.path.join(os.path.dirname(__file__), ".."))
DEFAULT_BOOK = os.environ.get("BOOK_XLSX", os.path.join(os.path.expanduser("~"), "Downloads", "Book.xlsx"))
OUT_YAML = os.path.join(ROOT, "config", "metrics_registry.yaml")
OUT_MD = os.path.join(ROOT, "docs", "METRICS_REGISTRY_SUMMARY.md")

TECHNIQUE_CODES = {
    ("Structural Analysis", "Cyclomatic Complexity"): "SA",
    ("Readability / Maintainability", "Cognitive Complexity"): "RM",
    ("Code Quality Auditing", "Code Duplication"): "CQ",
    ("Static Code Analysis", "Lint / Rule Violations"): "LR",
    ("Security White-box Testing", "Static Vulnerabilities (SAST)"): "SX",
    ("Security White-box Testing", "Dependency Risk (SCA)"): "DR",
    ("Control Flow Testing", "Statement Coverage"): "ST",
    ("Control Flow Testing", "Branch Coverage"): "BR",
    ("Control Flow Testing", "Path Coverage"): "PC",
    ("Mutation Testing", "Mutation Score"): "MU",
    ("Test Regression/Coverage Analysis", "Coverage Delta"): "CD",
    ("Data Flow Testing", "All Definition Coverage"): "DF",
    ("Data Flow Testing", "All Uses Coverage"): "DU",
    ("Development Process Analysis", "Code Churn"): "DP",
}

SA_METRIC_CODES = {
    "Execution Path Integrity": "EPI",
    "Decision Outcome Verification": "DOV",
    "Logical Sub-expression Validation": "LSV",
    "Total Logical Combinatorial Coverage": "TLCC",
    "Technical Debt Impact": "TDI",
    "QA Resource Allocation": "QRA",
}

SA_MODULE_KEYS = {
    "EPI": "execution_path_integrity",
    "DOV": "decision_coverage",
    "LSV": "condition_coverage",
    "TLCC": "logic_combinatorial",
    "TDI": "technical_debt",
    "QRA": "qa_prioritization",
}

LANG_COLS = {
    "python": (7, 8),
    "c": (11, 12),
    "cpp": (13, 14),
    "java": (15, 16),
    "csharp": (19, 20),
    "javascript": (21, 22),
    "typescript": (23, 24),
}

STOP = {"and", "or", "the", "a", "an", "of", "per", "for", "to", "in", "on", "all", "uses", "use"}


def _slug_words(name):
    words = re.findall(r"[A-Za-z0-9]+", name)
    return [w for w in words if w.lower() not in STOP]


def _acronym(name, max_len=5):
    words = _slug_words(name)
    if not words:
        return "MET"
    if len(words) == 1:
        return words[0][:max_len].upper()
    code = "".join(w[0] for w in words).upper()
    if len(code) < 3:
        code = "".join(w[:2] for w in words).upper()
    return code[:max_len]


def _module_key(metric_code, metric_name):
    return re.sub(r"[^a-z0-9]+", "_", metric_name.lower()).strip("_")[:48] or metric_code.lower()


def _clean_l5(l4, l5):
    l5s = (l5 or "").strip()
    if not l5s or "Retrieving data" in l5s:
        return (l4 or "").strip()
    return l5s


def parse_book(path):
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb["Sheet1"]
    techniques = {}
    metrics = []
    for r in range(6, ws.max_row + 1):
        l2 = ws.cell(r, 2).value
        l3 = ws.cell(r, 3).value
        l4 = ws.cell(r, 4).value
        l5_raw = ws.cell(r, 5).value
        if not l2 or not l3:
            continue
        l2, l3 = str(l2).strip(), str(l3).strip()
        l4 = str(l4).strip() if l4 else ""
        l5 = _clean_l5(l4, l5_raw)
        if not l5:
            continue
        key = (l2, l3)
        tech_code = TECHNIQUE_CODES.get(key)
        if not tech_code:
            raise ValueError("No technique_code for %s / %s (row %d)" % (l2, l3, r))
        if tech_code not in techniques:
            techniques[tech_code] = {
                "technique_code": tech_code,
                "l2": l2,
                "l3": l3,
                "report_group_label": l2,
                "metrics": [],
            }
        tools = {}
        for lang, (pc, sc) in LANG_COLS.items():
            pri, sec = ws.cell(r, pc).value, ws.cell(r, sc).value
            if pri or sec:
                tools[lang] = {
                    "primary": str(pri).strip() if pri else None,
                    "secondary": str(sec).strip() if sec else None,
                }
        metrics.append({
            "row": r,
            "technique_code": tech_code,
            "l4": l4,
            "l5": l5,
            "tools": tools,
        })

    # assign metric codes per technique group
    for tech_code, tech in techniques.items():
        used = set()
        group_rows = [m for m in metrics if m["technique_code"] == tech_code]
        for m in group_rows:
            if tech_code == "SA" and m["l5"] in SA_METRIC_CODES:
                code = SA_METRIC_CODES[m["l5"]]
            else:
                code = _acronym(m["l5"])
            base = code
            n = 2
            while code in used:
                code = "%s%d" % (base[:4], n)
                n += 1
            used.add(code)
            if tech_code == "SA":
                mod_key = SA_MODULE_KEYS.get(code, _module_key(code, m["l5"]))
            else:
                mod_key = _module_key(code, m["l5"])
            entry = {
                "metric_code": code,
                "module_key": mod_key,
                "l4_classification": m["l4"],
                "l5_metric": m["l5"],
                "taxonomy_classification": m["l4"],
                "tools": m["tools"],
                "source_row": m["row"],
            }
            tech["metrics"].append(entry)

    return {
        "version": "3.0",
        "source": os.path.basename(path),
        "branch_types": ["Bug", "BugFX", "TCC", "CC"],
        "techniques": list(techniques.values()),
    }


def write_outputs(data):
    os.makedirs(os.path.dirname(OUT_YAML), exist_ok=True)
    os.makedirs(os.path.dirname(OUT_MD), exist_ok=True)
    if yaml is None:
        raise RuntimeError("PyYAML required: pip install pyyaml")
    with open(OUT_YAML, "w", encoding="utf-8") as fh:
        yaml.dump(data, fh, default_flow_style=False, sort_keys=False, allow_unicode=True)
    lines = [
        "# Metrics registry summary",
        "",
        "Parsed from **%s** — %d technique groups, %d L5 metrics." % (
            data["source"],
            len(data["techniques"]),
            sum(len(t["metrics"]) for t in data["techniques"]),
        ),
        "",
        "| technique | metric | L5 metric | Python primary tool |",
        "|-----------|--------|-----------|---------------------|",
    ]
    for tech in data["techniques"]:
        for m in tech["metrics"]:
            py = (m.get("tools") or {}).get("python", {})
            pri = py.get("primary") or "—"
            lines.append("| %s | %s | %s | %s |" % (
                tech["technique_code"], m["metric_code"], m["l5_metric"], pri))
    with open(OUT_MD, "w", encoding="utf-8") as fh:
        fh.write("\n".join(lines) + "\n")


def main():
    path = sys.argv[1] if len(sys.argv) > 1 else DEFAULT_BOOK
    if not os.path.isfile(path):
        print("ERROR: Book.xlsx not found at %s" % path, file=sys.stderr)
        return 1
    data = parse_book(path)
    write_outputs(data)
    n_metrics = sum(len(t["metrics"]) for t in data["techniques"])
    print("Wrote %s (%d techniques, %d metrics)" % (OUT_YAML, len(data["techniques"]), n_metrics))
    print("Wrote %s" % OUT_MD)
    return 0


if __name__ == "__main__":
    sys.exit(main())
